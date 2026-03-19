import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, date, timedelta
import urllib.parse
import requests
import zipfile
import io
import os
import plotly.express as px
import openpyxl

# --- 1. CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Supervisión SEIN - Integral", layout="wide")
st.title("⚡ Dashboard de Supervisión - Mantenimientos y Despacho")
st.markdown("Fiscalización Dinámica, Asignación F/S y Análisis de Reservas Operativas (Gas y Diésel)")

MESES = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
    5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
    9: "SETIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
}

MES_TXT = [
    "ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO",
    "JULIO","AGOSTO","SETIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"
]

ARCHIVO_POTENCIAS_LOCAL = "potencias_historicas.csv"
ARCHIVO_GITHUB_SEMILLA = "pOTENCIAS.csv"
rdo_letras = list("ABCDEFGHIJ")

# --- DICCIONARIOS DE UNIDADES TÉRMICAS ---
INFO_UNIDADES_GAS = {
    "138": {"potencia": 89.51268, "unidad": "TG1", "central": "AGUAYTIA"},
    "139": {"potencia": 89.85903, "unidad": "TG2", "central": "AGUAYTIA"},
    "197": {"potencia": 182, "unidad": "TG1", "central": "KALLPA"},
    "203": {"potencia": 202, "unidad": "TG2", "central": "KALLPA"},
    "204": {"potencia": 205, "unidad": "TG3", "central": "KALLPA"},
    "230": {"potencia": 294, "unidad": "TV", "central": "KALLPA"},
    "209": {"potencia": 182.1, "unidad": "TG1", "central": "LAS FLORES"},
    "56677": {"potencia": 114, "unidad": "TV", "central": "LAS FLORES"},
    "252": {"potencia": 182, "unidad": "TG1", "central": "FENIX"},
    "249": {"potencia": 182, "unidad": "TG2", "central": "FENIX"},
    "250": {"potencia": 192, "unidad": "TV", "central": "FENIX"},
    "288": {"potencia": 46.58, "unidad": "TG-6", "central": "MALACAS 1"},
    "115": {"potencia": 88.22, "unidad": "TG-4", "central": "MALACAS 2"},
    "194": {"potencia": 180, "unidad": "TG1", "central": "CHILCA 1"},
    "196": {"potencia": 180, "unidad": "TG2", "central": "CHILCA 1"},
    "207": {"potencia": 208, "unidad": "TG3", "central": "CHILCA 1"},
    "236": {"potencia": 292, "unidad": "TV", "central": "CHILCA 1"},
    "795": {"potencia": 73.6, "unidad": "TG", "central": "CHILCA 2"},
    "285": {"potencia": 37.34, "unidad": "TV", "central": "CHILCA 2"},
    "110": {"potencia": 118.15, "unidad": "WTG-7", "central": "SANTA ROSA"},
    "113": {"potencia": 170, "unidad": "TG3", "central": "VENTANILLA"},
    "114": {"potencia": 170, "unidad": "TG4", "central": "VENTANILLA"},
    "193": {"potencia": 192, "unidad": "TV", "central": "VENTANILLA"},
    "208": {"potencia": 191.19621, "unidad": "TG8", "central": "SANTA ROSA"},
    "248": {"potencia": 210, "unidad": "TG1", "central": "OLLEROS"},
    "2159": {"potencia": 123, "unidad": "TV", "central": "OLLEROS"},
    "109": {"potencia": 52.41851, "unidad": "UTI-5", "central": "SANTA ROSA"},
    "144": {"potencia": 53.71012, "unidad": "UTI-6", "central": "SANTA ROSA"}
}

INFO_UNIDADES_DIESEL = {
    "42667": {"potencia": 8.82587, "unidad": "MAK1", "central": "TUMBES"},
    "42668": {"potencia": 8.51669, "unidad": "MAK2", "central": "TUMBES"},
    "142": {"potencia": 18.14123, "unidad": "TV1", "central": "SHOUGESA"},
    "143": {"potencia": 19.19196, "unidad": "TV2", "central": "SHOUGESA"},
    "125": {"potencia": 24.42536, "unidad": "TV3", "central": "SHOUGESA"},
    "187": {"potencia": 1.06127, "unidad": "CUMMINS", "central": "SHOUGESA"},
    "157": {"potencia": 11.41313, "unidad": "TG", "central": "CHILINA"},
    "156": {"potencia": 23.88915, "unidad": "CENTRAL", "central": "MOLLENDO"},
    "155": {"potencia": 10.22284, "unidad": "SULZ 1,2", "central": "CHILINA"},
    "758": {"potencia": 175.92342, "unidad": "CENTRAL", "central": "RECKA"},
    "263": {"potencia": 213.71525, "unidad": "TG1", "central": "RF ETEN"},
    "265": {"potencia": 7.84668, "unidad": "TG2", "central": "RF ETEN"},
    "240": {"potencia": 164.44227, "unidad": "TG1", "central": "RF ILO"},
    "241": {"potencia": 165.47058, "unidad": "TG2", "central": "RF ILO"},
    "242": {"potencia": 165.59853, "unidad": "TG3", "central": "CT NEPI"},
    "995": {"potencia": 206.49456, "unidad": "TG41", "central": "CT NEPI"},
    "996": {"potencia": 206.11549, "unidad": "TG42", "central": "CT NEPI"},
    "997": {"potencia": 205.99003, "unidad": "TG43", "central": "CT NEPI"},
    "926": {"potencia": 17.41583, "unidad": "CENTRAL", "central": "RF PTO MALDONADO"},
    "924": {"potencia": 44.05367, "unidad": "CENTRAL", "central": "RF PUCALLPA"},
    "786": {"potencia": 180.55348, "unidad": "TG1", "central": "PTO BRAVO"},
    "787": {"potencia": 180.52451, "unidad": "TG2", "central": "PTO BRAVO"},
    "788": {"potencia": 181.33356, "unidad": "TG3", "central": "PTO BRAVO"},
    "789": {"potencia": 181.02912, "unidad": "TG4", "central": "PTO BRAVO"}
}

CODIGOS_PERMITIDOS_TOTAL = list(INFO_UNIDADES_GAS.keys()) + list(INFO_UNIDADES_DIESEL.keys())

COLOR_MAP = {
    "Hidráulica": "#3498db",
    "Eólica": "#2ecc71",
    "Solar": "#f1c40f",
    "Biogás+Biomasa+Nafta+Flexigas": "#800080",
    "Gas de Camisea": "#006400",
    "Gas del Norte+Gas de la Selva": "#90EE90",
    "Residual+Diésel D2": "#FF0000",
    "Otros": "#95a5a6"
}

def get_stable_color_map(unidades):
    colores_base = px.colors.qualitative.Alphabet + px.colors.qualitative.Dark24
    return {uni: colores_base[i % len(colores_base)] for i, uni in enumerate(sorted(unidades))}

# --- 2. FUNCIONES BASE Y CLASIFICADORES ---
def normalizar_texto(serie):
    return serie.astype(str).str.strip().str.upper().str.normalize('NFKD').str.encode('ascii', errors='ignore').str.decode('utf-8')

def determinar_sector(row):
    tipo = str(row.get('Tipo_Equipo', '')).strip().upper()
    equipo = str(row.get('Equipo', '')).strip().upper()
    if tipo.startswith('G'): return 'GENERACIÓN'
    if tipo.startswith('T') or tipo.startswith('L'): return 'TRANSMISIÓN'
    if equipo.startswith('L-') or equipo.startswith('TR') or equipo.startswith('AT') or equipo.startswith('SE '): return 'TRANSMISIÓN'
    if equipo.startswith('G-') or equipo.startswith('TV') or equipo.startswith('TG') or equipo.startswith('CH '): return 'GENERACIÓN'
    return 'OTROS'

def clasificar_tecnologia_yupana(nombre_central):
    nombre = str(nombre_central).upper().strip()
    if any(kw in nombre for kw in ["(HID)", "CH ", "C.H.", "RESTITUCION", "MANTARO", "HUINCO", "CHARCANI", "MACHUPICCHU", "GALLITO", "CAÑON", "SAN GABAN", "YUNCAN", "CHAGLLA", "CERRO DEL AGUILA", "QUELCAYA", "CANA BRAVA"]): return "Hidráulica"
    if any(kw in nombre for kw in ["(EOL)", "CE ", "WAYRA", "TRES HERMANAS", "MARCONA", "CUPISNIQUE", "TALARA"]): return "Eólica"
    if any(kw in nombre for kw in ["(SOL)", "CS ", "RUBI", "INTIPAMPA", "PANAMERICANA", "MOQUEGUA FV"]): return "Solar"
    biomasa_kws = ["PARAMONGA", "JACINTO", "HUAYCOLORO", "GRINGA", "MAPLE", "FLEXIGAS", "NAFTA", "LUREN", "BIOMASA", "BIOGAS", "AGROAURORA", "CAHUAPANAS", "SUPE", "LAREDO", "PETRAMAS", "DOÑA CATALINA", "PORTILLO", "CASA GRANDE", "AGROOLMOS", "CALLAO", "REFTALARA"]
    if any(kw in nombre for kw in biomasa_kws): return "Biogás+Biomasa+Nafta+Flexigas"
    diesel_kws = ["D2", "R6", "RESIDUAL", "DIESEL", "ILO", "MOLLENDO", "RECKA", "INDEPENDENCIA", "SAMANCO", "TARAPOTO", "IQUITOS", "YURIMAGUAS", "PUERTO MALDONADO", "BELLAVISTA", "PEDRO RUIZ", "ETEN", "PIURA D", "CALANA", "ELOR", "SHCUMMINS", "SNTV", "NEPI", "PUERTO BRAVO", "NODO"]
    if any(kw in nombre for kw in diesel_kws): return "Residual+Diésel D2"
    duales_gas_kws = ["FENIX", "KALLPA", "CHILCA", "VENTANILLA", "LAS FLORES", "SANTO DOMINGO", "MALACAS", "TALLANCA", "AGUAYTIA", "TERMOSELVA"]
    if any(ex in nombre for ex in duales_gas_kws):
        if any(kw in nombre for kw in ["MALACAS", "TALLANCA", "AGUAYTIA", "TERMOSELVA"]): return "Gas del Norte+Gas de la Selva"
        return "Gas de Camisea"
    gas_norte_kws = ["PUCALLPA", "ZORRITOS", "PARIÑAS", "EEEP", "ENEL PIURA", "PIURA G", "NUEVA ZORRITOS", "AGE", "MAL2", "TABLAZO"]
    if any(kw in nombre for kw in gas_norte_kws): return "Gas del Norte+Gas de la Selva"
    return "Gas de Camisea" if "(TER)" in nombre else "Otros"

def cargar_potencias_guardadas():
    if os.path.exists(ARCHIVO_POTENCIAS_LOCAL):
        df = pd.read_csv(ARCHIVO_POTENCIAS_LOCAL)
    elif os.path.exists(ARCHIVO_GITHUB_SEMILLA):
        df = pd.read_csv(ARCHIVO_GITHUB_SEMILLA)
        if 'Empresa' not in df.columns: df['Empresa'] = 'NO ESPECIFICADO'
    else:
        df = pd.DataFrame(columns=['Empresa', 'Central/Ubicacion', 'Equipo', 'Potencia_Indisponible_MW'])
    return df.drop_duplicates(subset=['Central/Ubicacion', 'Equipo'], keep='last')

def guardar_potencias_asignadas(df_nuevas):
    df_historico = cargar_potencias_guardadas()
    if not df_historico.empty:
        df_final = pd.concat([df_nuevas, df_historico]).drop_duplicates(subset=['Central/Ubicacion', 'Equipo'], keep='first')
    else:
        df_final = df_nuevas
    df_final.to_csv(ARCHIVO_POTENCIAS_LOCAL, index=False)

# --- 3. MOTOR DE EXTRACCIÓN (MANTENIMIENTO Y EXCEL PDO/RDO) ---
def generar_url_coes_mant(fecha):
    año = fecha.strftime("%Y")
    mes_num = fecha.strftime("%m")
    dia = fecha.strftime("%d")
    mes_mayus = MESES[fecha.month]
    fecha_str = fecha.strftime("%Y%m%d")
    path_prog = f"Operación/Programa de Mantenimiento/Programa Diario/{año}/{mes_num}_{mes_mayus}/Día {dia}/Anexo1_Intervenciones_{fecha_str}.zip"
    return f"https://www.coes.org.pe/portal/browser/download?url={urllib.parse.quote(path_prog)}"

@st.cache_data(show_spinner=False, ttl=3600)
def extraer_anexo_osinergmin(fecha):
    url = generar_url_coes_mant(fecha)
    headers = {'User-Agent': 'Mozilla/5.0'}
    columnas_estandar = ['Empresa', 'Ubicacion', 'Equipo', 'Inicio', 'Fin', 'Descripcion', 'MW_Indisponibles', 'Es_Programado', 'Disponibilidad_Equipo', 'Ocasiona_Interrupciones', 'Tipo_Mantenimiento', 'Codigo_Equipo', 'Tipo_Equipo']
    df_prog = pd.DataFrame()
    try:
        res = requests.get(url, headers=headers, timeout=15)
        if res.status_code == 200 and res.content[:4] == b'PK\x03\x04':
            with zipfile.ZipFile(io.BytesIO(res.content)) as z:
                archivo_objetivo = next((f for f in z.namelist() if 'Anexo1_Intervenciones_(Osinergmin)' in f and not f.startswith('~$') and '__MACOSX' not in f), None)
                if archivo_objetivo:
                    with z.open(archivo_objetivo) as f:
                        excel_file = pd.ExcelFile(io.BytesIO(f.read()), engine='xlrd' if archivo_objetivo.endswith('.xls') else 'openpyxl')
                        hoja_prog = next((h for h in excel_file.sheet_names if 'PROGRAMADO' in h.upper() or 'ANEXO' in h.upper()), excel_file.sheet_names[0])
                        df_prog = pd.read_excel(excel_file, sheet_name=hoja_prog, skiprows=8, usecols="B:N", names=columnas_estandar)
                        df_prog = df_prog.dropna(subset=['Empresa', 'Equipo'], how='all')
                        df_prog = df_prog[~df_prog['Empresa'].astype(str).str.contains('TOTAL|NOTA|ELABORADO|FUENTE', case=False, na=False)]
                        df_prog['Fecha_Reporte'] = fecha
    except: pass
    return df_prog

@st.cache_data(show_spinner=False, ttl=3600)
def extraer_datos_dia_excel(f):
    y, m, d = f.strftime("%Y"), f.strftime("%m"), f.strftime("%d")
    M = MES_TXT[f.month-1]
    ddmm = f"{d}{m}"
    fecha_str = f"{y}{m}{d}"
    
    headers = {'User-Agent': 'Mozilla/5.0'}
    
    path_pdo_principal = f"Operación/Programa de Operación/Programa Diario/{y}/{m}_{M}/Día {d}/Anexo1_Despacho_{fecha_str}.xlsx"
    
    urls_a_intentar = {
        "PDO": [
            f"https://www.coes.org.pe/portal/browser/download?url={urllib.parse.quote(path_pdo_principal)}",
            f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FPrograma%20Diario%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FProg_{ddmm}.xlsx",
            f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FPrograma%20Diario%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FProg%20{ddmm}.xlsx",
            f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FPrograma%20Diario%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FPROG_{ddmm}.xlsx"
        ]
    }
    for l in rdo_letras:
        urls_a_intentar[f"RDO_{l}"] = [
            f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FReprograma%20Diario%20Operaci%C3%B3n%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FReprog%20{ddmm}{l}%2FReprog_{ddmm}{l}.xlsx",
            f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FReprograma%20Diario%20Operaci%C3%B3n%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FReprog%20{ddmm}%20{l}%2FReprog_{ddmm}{l}.xlsx"
        ]

    datos_dia = {"Dataframes": {}}
    for nombre, lista_enlaces in urls_a_intentar.items():
        exito = False
        for enlace in lista_enlaces:
            if exito: break
            try:
                r = requests.get(enlace, headers=headers, timeout=20)
                if r.status_code == 200 and len(r.content) > 5000:
                    wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
                    ws = wb.worksheets[0]
                    
                    df_headers = pd.read_excel(io.BytesIO(r.content), sheet_name=0, header=None, skiprows=4, nrows=2, usecols="B:CU")
                    codes = df_headers.iloc[0].values
                    
                    df = pd.read_excel(io.BytesIO(r.content), sheet_name=0, header=5, usecols="B:CU")
                    df.rename(columns={df.columns[0]: 'Hora'}, inplace=True)
                    
                    cutoff_idx = df.index[df['Hora'].astype(str).str.strip().str.upper() == 'MWH'].tolist()
                    if cutoff_idx:
                        df = df.iloc[:cutoff_idx[0]]
                    
                    cols_validas = ['Hora']
                    name_to_code = {}
                    
                    for idx, c in enumerate(codes):
                        if idx == 0: continue
                        if idx < len(df.columns):
                            c_str = str(c).replace(".0", "").strip()
                            if c_str in CODIGOS_PERMITIDOS_TOTAL:
                                col_name_df = df.columns[idx]
                                cols_validas.append(col_name_df)
                                name_to_code[col_name_df] = c_str
                                
                    df = df[cols_validas]
                    df.rename(columns=name_to_code, inplace=True)
                    
                    start_idx = max(0, 48 - len(df))
                    if len(df) < 48:
                        pad_df = pd.DataFrame(0.0, index=range(start_idx), columns=df.columns)
                        df_padded = pd.concat([pad_df, df], ignore_index=True)
                    else:
                        df_padded = df.iloc[:48].copy()
                        df_padded.reset_index(drop=True, inplace=True)
                        
                    datos_dia["Dataframes"][nombre] = {"EXCEL": df_padded}
                    datos_dia["Dataframes"][f"START_IDX_{nombre}"] = start_idx
                    datos_dia["Dataframes"]["NAME_TO_CODE"] = name_to_code
                    
                    hora_inicio = "N/A"
                    motivo = "No se encontró justificación."
                    if "RDO" in nombre:
                        val_b7 = ws.cell(row=7, column=2).value
                        if val_b7 is not None:
                            if isinstance(val_b7, datetime) or hasattr(val_b7, 'strftime'):
                                hora_inicio = val_b7.strftime("%H:%M")
                            else:
                                val_str = str(val_b7).strip()
                                if len(val_str) >= 5 and ":" in val_str: 
                                    hora_inicio = val_str[:5]
                                elif ":" in val_str: 
                                    hora_inicio = val_str
                                    
                        for row in range(1, min(ws.max_row, 150) + 1):
                            val_c = str(ws.cell(row=row, column=3).value).upper()
                            if "MOTIVO" in val_c:
                                val_mot = ws.cell(row=row+1, column=4).value
                                if val_mot: motivo = str(val_mot).strip()
                                break
                                
                    datos_dia["Dataframes"][f"MOTIVO_{nombre}"] = motivo
                    datos_dia["Dataframes"][f"HORA_{nombre}"] = hora_inicio
                    exito = True
            except: pass
        if not exito and "RDO" in nombre: break
    return datos_dia

# --- 4. FUNCIÓN GRAFICADORA DINÁMICA ---
def crear_grafica_dinamica(df_plot, marcadores=None, tipo_grafico="area", paleta_distinta=False):
    df_plot = df_plot.copy().fillna(0)
    num_cols = [c for c in df_plot.columns if c != 'Hora']
    df_plot[num_cols] = df_plot[num_cols].apply(pd.to_numeric, errors='coerce').fillna(0).round(2)
    
    if tipo_grafico == "barra":
        df_plot['Fecha'] = (df_plot['Hora'] - pd.Timedelta(minutes=1)).dt.strftime('%d/%m/%Y')
        df_daily = df_plot.groupby('Fecha')[num_cols].sum() / 2.0
        df_daily = df_daily.reset_index()
        df_daily['TOTAL_GRAFICA'] = df_daily[num_cols].sum(axis=1).round(2)
        
        totales = df_daily.drop(columns=['Fecha', 'TOTAL_GRAFICA']).sum()
        orden_columnas = totales.sort_values(ascending=False).index.tolist()
            
        cols_mantener = ['Fecha', 'TOTAL_GRAFICA'] + orden_columnas
        df_melt = df_daily[cols_mantener].melt(id_vars=['Fecha', 'TOTAL_GRAFICA'], var_name='Unidad Generadora', value_name='Energía_MWh')
        
        if paleta_distinta:
            color_map_dinamico = get_stable_color_map(df_melt['Unidad Generadora'].unique())
        else:
            color_map_dinamico = {uni: COLOR_MAP.get(clasificar_tecnologia_yupana(uni), "#95a5a6") for uni in df_melt['Unidad Generadora'].unique()}
        
        fig = px.bar(df_melt, x="Fecha", y="Energía_MWh", color="Unidad Generadora", labels={"Energía_MWh": "Energía Diaria (MWh)"}, color_discrete_map=color_map_dinamico)
        fig.add_scatter(x=df_daily['Fecha'], y=df_daily['TOTAL_GRAFICA'], mode='markers+text', text=df_daily['TOTAL_GRAFICA'].apply(lambda x: f"{x:,.1f} MWh"), textposition="top center", marker=dict(color='rgba(0,0,0,0)'), name='<b>∑ TOTAL</b>', showlegend=False)
        for trace in fig.data:
            if trace.name and 'TOTAL' in trace.name: pass
            else: trace.hovertemplate = "%{y:,.2f} MWh<br>%{x}"
        fig.update_layout(hovermode="x unified", height=650, margin=dict(t=170, b=50, l=60, r=50), xaxis_title="Fecha Operativa")
        return fig
    else:
        df_plot['TOTAL_GRAFICA'] = df_plot[num_cols].sum(axis=1).round(2)
        totales_por_unidad = df_plot.drop(columns=['Hora', 'TOTAL_GRAFICA']).sum()
        orden_columnas = totales_por_unidad.sort_values(ascending=False).index.tolist()
        
        cols_mantener = ['Hora', 'TOTAL_GRAFICA'] + orden_columnas
        df_melt = df_plot[cols_mantener].melt(id_vars=['Hora', 'TOTAL_GRAFICA'], var_name='Unidad Generadora', value_name='Potencia_MW')
        
        if paleta_distinta:
            color_map_dinamico = get_stable_color_map(df_melt['Unidad Generadora'].unique())
        else:
            color_map_dinamico = {uni: COLOR_MAP.get(clasificar_tecnologia_yupana(uni), "#95a5a6") for uni in df_melt['Unidad Generadora'].unique()}
        
        fig = px.area(df_melt, x="Hora", y="Potencia_MW", color="Unidad Generadora", labels={"Potencia_MW": "Potencia Activa (MW)"}, color_discrete_map=color_map_dinamico)
        fig.update_xaxes(tickformat="%d/%m %H:%M", tickangle=45)
        fig.update_traces(line=dict(width=0)) 
        
        fig.add_scatter(x=df_plot['Hora'], y=df_plot['TOTAL_GRAFICA'], mode='lines', line=dict(width=0, color='rgba(0,0,0,0)'), name='<b>∑ TOTAL</b>', showlegend=False)
        
        if marcadores:
            for ts, texto in marcadores:
                fig.add_vline(x=ts, line_width=1.5, line_dash="dash", line_color="rgba(0,0,0,0.5)")
                texto_con_hora = f"{texto.replace('(', '').replace(')', '')} {ts.strftime('%H:%M')}"
                align = "left" if ts.hour == 0 and ts.minute == 30 else "center"
                fig.add_annotation(x=ts, y=1.08, yref="paper", text=f"<b>{texto_con_hora}</b>", showarrow=False, font=dict(size=10, color="white"), bgcolor="#e74c3c", textangle=-90, yanchor="bottom", xanchor=align)
                
        fig.update_layout(hovermode="x unified", height=650, margin=dict(t=170, b=50, l=60, r=50))
        return fig

# --- 5. EJECUCIÓN PRINCIPAL ---
st.sidebar.header("Parámetros de Fiscalización")
hoy = date.today()
rango_fechas = st.sidebar.date_input("Rango Operativo", value=(hoy - timedelta(days=2), hoy))

if st.sidebar.button("Extraer Consolidado Integral", type="primary"):
    if isinstance(rango_fechas, tuple) and len(rango_fechas) == 2:
        with st.spinner("Compilando Mantenimientos y Motor Excel PDO/RDO (Analizando Gas y Diésel)..."):
            rango_dias = pd.date_range(rango_fechas[0], rango_fechas[1])
            dfs_mantenimiento = []
            datos_yupana_completos = {}
            
            barra = st.progress(0)
            for i, d in enumerate(rango_dias):
                df_dia_mant = extraer_anexo_osinergmin(d.date())
                if not df_dia_mant.empty: dfs_mantenimiento.append(df_dia_mant)
                
                datos_yupana_completos[d.date()] = extraer_datos_dia_excel(d.date())
                barra.progress((i + 1) / len(rango_dias))
            barra.empty()
            
            df_mant_final = pd.DataFrame()
            if dfs_mantenimiento:
                df_mant_final = pd.concat(dfs_mantenimiento, ignore_index=True)
                df_mant_final['Empresa'] = normalizar_texto(df_mant_final['Empresa'])
                df_mant_final['Central/Ubicacion'] = normalizar_texto(df_mant_final['Ubicacion'].fillna('-'))
                df_mant_final['Equipo'] = normalizar_texto(df_mant_final['Equipo'])
                df_mant_final['Tipo_Mantenimiento'] = df_mant_final['Tipo_Mantenimiento'].fillna('NO ESPECIFICADO').astype(str).str.strip().str.upper()
                df_mant_final['Disponibilidad_Equipo'] = df_mant_final['Disponibilidad_Equipo'].fillna('NO ESPECIFICADO').astype(str).str.strip().str.upper()
                df_mant_final['Sector'] = df_mant_final.apply(determinar_sector, axis=1)
                df_mant_final['MW_Indisponibles'] = pd.to_numeric(df_mant_final['MW_Indisponibles'], errors='coerce').fillna(0)
                df_mant_final['Inicio_DT'] = pd.to_datetime(df_mant_final['Inicio'], errors='coerce', dayfirst=True)
                df_mant_final['Fin_DT'] = pd.to_datetime(df_mant_final['Fin'], errors='coerce', dayfirst=True)
                df_mant_final['Horas_Maniobra'] = np.round((df_mant_final['Fin_DT'] - df_mant_final['Inicio_DT']).dt.total_seconds() / 3600, 2)
            
            st.session_state['df_maestro'] = df_mant_final
            st.session_state['datos_yupana'] = datos_yupana_completos
    else:
        st.sidebar.error("Seleccione una fecha de inicio y fin válidas.")

# INICIALIZACIÓN DE GRÁFICAS PARA EL RESUMEN GLOBAL
fig_disp_gas, fig_res_gas, df_reserva_gas, df_total_gas = None, None, pd.DataFrame(), pd.DataFrame()
fig_disp_die, fig_res_die, df_reserva_die, df_total_d = None, None, pd.DataFrame(), pd.DataFrame()

if 'df_maestro' in st.session_state and 'datos_yupana' in st.session_state:
    df = st.session_state['df_maestro']
    data_yupana = st.session_state['datos_yupana']
    
    fechas_ordenadas = sorted(data_yupana.keys())
    active_prog_dict, ts_dict, dics_cache_dict, marcadores_globales = {}, {}, {}, []
    
    for f in fechas_ordenadas:
        df_dia_sel = data_yupana[f]["Dataframes"]
        progs = [p for p in ["PDO"] + [f"RDO_{l}" for l in rdo_letras] if p in df_dia_sel]
        if not progs: continue
        
        dics_cache = {p: df_dia_sel[p] for p in progs}
        active_prog = [progs[0]] * 48
        
        if len(progs) > 1:
            for p in progs[1:]:
                hora_inicio_str = df_dia_sel.get(f"HORA_{p}")
                inicio_idx = None
                if hora_inicio_str and hora_inicio_str != "N/A":
                    try:
                        partes = hora_inicio_str.split(":")
                        h = int(partes[0])
                        m = int(partes[1])
                        inicio_idx = h * 2 + (1 if m >= 30 else 0)
                    except: pass
                
                if inicio_idx is None: 
                    inicio_idx = df_dia_sel.get(f"START_IDX_{p}", 0)
                
                if inicio_idx is not None and 0 <= inicio_idx < 48:
                    for j in range(inicio_idx, 48): active_prog[j] = p
                        
        ts_dia = [datetime.combine(f, datetime.min.time()) + timedelta(minutes=30*(i+1)) for i in range(48)]
        p_actual = active_prog[0]
        marcadores_globales.append((ts_dia[0], p_actual))
        
        for i in range(1, 48):
            if active_prog[i] != p_actual:
                p_actual = active_prog[i]
                marcadores_globales.append((ts_dia[i-1], p_actual))
                
        active_prog_dict[f] = active_prog
        ts_dict[f] = ts_dia
        dics_cache_dict[f] = dics_cache

    st.markdown("### 🎛️ Filtros de Fiscalización Dinámicos")
    col1, col2, col3, col4, col5 = st.columns(5)
    
    def get_unique_safe(df_safe, col_name):
        if not df_safe.empty and col_name in df_safe.columns:
            return sorted(df_safe[col_name].dropna().unique())
        return []

    opciones_empresa = get_unique_safe(df, 'Empresa')
    opciones_central = get_unique_safe(df, 'Central/Ubicacion')
    opciones_sector = get_unique_safe(df, 'Sector')
    opciones_estado = get_unique_safe(df, 'Disponibilidad_Equipo')
    opciones_tipo = get_unique_safe(df, 'Tipo_Mantenimiento')

    sel_empresa = col1.multiselect("Concesionaria:", opciones_empresa, default=[])
    sel_central = col2.multiselect("Central/Ubicación:", opciones_central, default=[])
    sel_sector = col3.multiselect("Sector:", opciones_sector, default=['GENERACIÓN'] if 'GENERACIÓN' in opciones_sector else [])
    default_estado = ['F/S'] if 'F/S' in opciones_estado else []
    sel_estado = col4.multiselect("Disponibilidad (E/S - F/S):", opciones_estado, default=default_estado)
    sel_tipo = col5.multiselect("Tipo de Mantenimiento:", opciones_tipo, default=[])
    
    df_filtrado = df.copy()
    if not df_filtrado.empty:
        if sel_empresa: df_filtrado = df_filtrado[df_filtrado['Empresa'].isin(sel_empresa)]
        if sel_central: df_filtrado = df_filtrado[df_filtrado['Central/Ubicacion'].isin(sel_central)]
        if sel_sector: df_filtrado = df_filtrado[df_filtrado['Sector'].isin(sel_sector)]
        if sel_estado: df_filtrado = df_filtrado[df_filtrado['Disponibilidad_Equipo'].isin(sel_estado)]
        if sel_tipo: df_filtrado = df_filtrado[df_filtrado['Tipo_Mantenimiento'].isin(sel_tipo)]
        
        df_historico_pot = cargar_potencias_guardadas()
        if 'Central/Ubicacion' in df_filtrado.columns and 'Equipo' in df_filtrado.columns:
            df_filtrado = pd.merge(df_filtrado, df_historico_pot[['Central/Ubicacion', 'Equipo', 'Potencia_Indisponible_MW']], on=['Central/Ubicacion', 'Equipo'], how='left')
            df_filtrado['MW_Efectivos'] = df_filtrado['Potencia_Indisponible_MW'].fillna(df_filtrado['MW_Indisponibles']).fillna(0.0)

    # Base Global de F/S segura
    if not df_filtrado.empty and 'Sector' in df_filtrado.columns and 'Disponibilidad_Equipo' in df_filtrado.columns:
        df_gen_fs = df_filtrado[(df_filtrado['Sector'] == 'GENERACIÓN') & (df_filtrado['Disponibilidad_Equipo'] == 'F/S')].copy()
    else:
        df_gen_fs = pd.DataFrame()

    st.markdown("---")
    
    t_resumen_global, t_resumen_mant, t_indisponible, t_termico_gas, t_termico_diesel, t_motivos, t_datos = st.tabs([
        "🌟 1. Resumen Global (Reservas)",
        "📊 2. Resumen Mantenimientos", 
        "🔌 3. Indisponibilidad F/S", 
        "⚡ 4. Despacho Excel (Gas)",
        "🛢️ 5. Despacho Excel (Diésel)",
        "📋 6. Motivos RDO",
        "🗄️ 7. Trazabilidad de Datos"
    ])
    
    # ---------------------------------------------------------
    # PESTAÑA 1 (PARTE A): ASIGNACIÓN Y RESUMEN GLOBAL 
    # ---------------------------------------------------------
    with t_resumen_global:
        st.subheader("🌟 Resumen Global de Seguridad del Sistema")
        st.info("📖 **Guía de Asignación Operativa (PASO 1):**\n\n"
                "El cálculo del margen de reserva operativa consiste en restar el Despacho Real a la Potencia Efectiva disponible. Para que el cálculo sea 100% exacto, configura las unidades de la siguiente manera:\n"
                "- **Potencia Efectiva:** Ajusta el valor libremente. \n"
                "- **Vincular Mantenimiento COES:** Si seleccionas la máquina en esta lista, el motor buscará automáticamente *todos* los F/S reportados en el Anexo para esa unidad y anulará su reserva a 0 MW en esos intervalos.\n"
                "- **Fechas Manuales:** Introduce un inicio y fin (DD/MM/AAAA HH:mm) para forzar caídas de reserva no reportadas.\n"
                "*Al finalizar, desplázate hacia abajo para analizar las curvas de Reserva y Despacho.*")

        opciones_mant_gas = ["Ninguno"]
        opciones_mant_die = ["Ninguno"]
        if not df_gen_fs.empty:
            for _, row_fs in df_gen_fs.iterrows():
                equipo_str = str(row_fs['Equipo']).strip().upper()
                central_str = str(row_fs['Central/Ubicacion']).strip().upper()
                label_mant = f"{central_str} - {equipo_str}"
                if label_mant not in opciones_mant_gas: opciones_mant_gas.append(label_mant)
                if label_mant not in opciones_mant_die: opciones_mant_die.append(label_mant)

        # -- ASIGNACIÓN GAS --
        with st.expander("📝 1. Matriz de Asignación de Potencia y Mantenimientos (GAS)", expanded=False):
            df_potencias_gas = pd.DataFrame([
                {"Central": info["central"], "Unidad": info["unidad"], "Código": cod, 
                 "Potencia Efectiva (MW)": info["potencia"], "Vincular Mantenimiento (COES)": "Ninguno",
                 "Inicio Manual": None, "Fin Manual": None}
                for cod, info in INFO_UNIDADES_GAS.items()
            ])
            df_editado_gas = st.data_editor(
                df_potencias_gas,
                column_config={
                    "Central": st.column_config.TextColumn("Central", disabled=True),
                    "Unidad": st.column_config.TextColumn("Unidad", disabled=True),
                    "Código": st.column_config.TextColumn("Código", disabled=True),
                    "Potencia Efectiva (MW)": st.column_config.NumberColumn("Potencia Efectiva (MW)", min_value=0.0, format="%.2f"),
                    "Vincular Mantenimiento (COES)": st.column_config.SelectboxColumn("Vincular Mantenimiento COES", options=opciones_mant_gas),
                    "Inicio Manual": st.column_config.DatetimeColumn("Inicio Manual (DD/MM/AAAA HH:mm)", format="DD/MM/YYYY HH:mm"),
                    "Fin Manual": st.column_config.DatetimeColumn("Fin Manual (DD/MM/AAAA HH:mm)", format="DD/MM/YYYY HH:mm")
                }, use_container_width=True, hide_index=True, key="ed_gas_global"
            )
            
        dict_pot_editada = {}
        dict_mant_vinculado = {}
        dict_mant_manual = {}
        for _, row_ed in df_editado_gas.iterrows():
            cod_ed = str(row_ed["Código"])
            dict_pot_editada[cod_ed] = float(row_ed["Potencia Efectiva (MW)"])
            if row_ed["Vincular Mantenimiento (COES)"] != "Ninguno":
                dict_mant_vinculado[cod_ed] = row_ed["Vincular Mantenimiento (COES)"]
            ini_man = pd.to_datetime(row_ed["Inicio Manual"], errors='coerce')
            fin_man = pd.to_datetime(row_ed["Fin Manual"], errors='coerce')
            if pd.notna(ini_man) and pd.notna(fin_man):
                dict_mant_manual[cod_ed] = (ini_man, fin_man)

        # -- ASIGNACIÓN DIÉSEL --
        with st.expander("📝 2. Matriz de Asignación de Potencia y Mantenimientos (DIÉSEL)", expanded=False):
            df_potencias_d = pd.DataFrame([
                {"Central": info["central"], "Unidad": info["unidad"], "Código": cod, 
                 "Potencia Efectiva (MW)": info["potencia"], "Vincular Mantenimiento (COES)": "Ninguno",
                 "Inicio Manual": None, "Fin Manual": None}
                for cod, info in INFO_UNIDADES_DIESEL.items()
            ])
            df_editado_d = st.data_editor(
                df_potencias_d,
                column_config={
                    "Central": st.column_config.TextColumn("Central", disabled=True),
                    "Unidad": st.column_config.TextColumn("Unidad", disabled=True),
                    "Código": st.column_config.TextColumn("Código", disabled=True),
                    "Potencia Efectiva (MW)": st.column_config.NumberColumn("Potencia Efectiva (MW)", min_value=0.0, format="%.2f"),
                    "Vincular Mantenimiento (COES)": st.column_config.SelectboxColumn("Vincular Mantenimiento COES", options=opciones_mant_die),
                    "Inicio Manual": st.column_config.DatetimeColumn("Inicio Manual (DD/MM/AAAA HH:mm)", format="DD/MM/YYYY HH:mm"),
                    "Fin Manual": st.column_config.DatetimeColumn("Fin Manual (DD/MM/AAAA HH:mm)", format="DD/MM/YYYY HH:mm")
                }, use_container_width=True, hide_index=True, key="ed_die_global"
            )
            
        dict_pot_editada_d = {}
        dict_mant_vinc_d = {}
        dict_mant_man_d = {}
        for _, row_ed in df_editado_d.iterrows():
            cod_ed = str(row_ed["Código"])
            dict_pot_editada_d[cod_ed] = float(row_ed["Potencia Efectiva (MW)"])
            if row_ed["Vincular Mantenimiento (COES)"] != "Ninguno":
                dict_mant_vinc_d[cod_ed] = row_ed["Vincular Mantenimiento (COES)"]
            ini_man = pd.to_datetime(row_ed["Inicio Manual"], errors='coerce')
            fin_man = pd.to_datetime(row_ed["Fin Manual"], errors='coerce')
            if pd.notna(ini_man) and pd.notna(fin_man):
                dict_mant_man_d[cod_ed] = (ini_man, fin_man)

    # ---------------------------------------------------------
    # PESTAÑA 2: RESUMEN MANTENIMIENTOS
    # ---------------------------------------------------------
    with t_resumen_mant:
        if not df_filtrado.empty and 'MW_Efectivos' in df_filtrado.columns:
            pico_maximo_mw = 0.0
            if not df_gen_fs.empty:
                df_fs = df_gen_fs.dropna(subset=['Inicio_DT', 'Fin_DT']).copy()
                df_fs['Central_Equipo'] = df_fs['Central/Ubicacion'] + " - " + df_fs['Equipo']
                min_dt, max_dt = df_fs['Inicio_DT'].min(), df_fs['Fin_DT'].max()
                if min_dt < max_dt:
                    time_grid = pd.date_range(start=min_dt, end=max_dt, freq='h')
                    for t in time_grid:
                        mask = (df_fs['Inicio_DT'] <= t) & (df_fs['Fin_DT'] > t)
                        df_t_unique = df_fs.loc[mask].drop_duplicates(subset=['Central_Equipo'])
                        centrales = df_t_unique[df_t_unique['Equipo'] == 'CENTRAL'][['Empresa', 'Central/Ubicacion']].drop_duplicates()
                        for _, row_c in centrales.iterrows():
                            condicion = (df_t_unique['Empresa'] == row_c['Empresa']) & (df_t_unique['Central/Ubicacion'] == row_c['Central/Ubicacion']) & (df_t_unique['Equipo'] != 'CENTRAL')
                            df_t_unique = df_t_unique[~condicion]
                        mw_hora = df_t_unique['MW_Efectivos'].sum()
                        if mw_hora > pico_maximo_mw: pico_maximo_mw = mw_hora
            
            energia_total = (df_filtrado['MW_Efectivos'] * df_filtrado['Horas_Maniobra']).sum()
            
            kpi1, kpi2, kpi3, kpi4 = st.columns(4)
            kpi1.metric("Maniobras", f"{len(df_filtrado)}")
            kpi2.metric("Energía Total Involucrada", f"{energia_total:,.1f} MWh")
            kpi3.metric("Máxima Indisp. Simultánea", f"{pico_maximo_mw:,.2f} MW")
            kpi4.metric("Tiempo Total de Intervención", f"{df_filtrado['Horas_Maniobra'].sum():,.1f} h")
            
            c_g1, c_g2 = st.columns(2)
            with c_g1:
                df_empresa = df_filtrado.groupby('Empresa')['MW_Efectivos'].sum().reset_index().sort_values('MW_Efectivos', ascending=False).head(10)
                fig_bar_emp = px.bar(df_empresa, x='MW_Efectivos', y='Empresa', orientation='h', title="Top 10 Empresas por Capacidad Intervenida (MW Efectivo)", text_auto='.1f')
                fig_bar_emp.update_layout(yaxis={'categoryorder':'total ascending'})
                st.plotly_chart(fig_bar_emp, use_container_width=True, key="graf_mant_emp")
                
            with c_g2:
                df_sector = df_filtrado.groupby('Sector')['Horas_Maniobra'].count().reset_index()
                fig_bar_sec = px.bar(df_sector, x='Sector', y='Horas_Maniobra', title="Cantidad de Maniobras por Sector", text_auto=True, color='Sector')
                st.plotly_chart(fig_bar_sec, use_container_width=True, key="graf_mant_sec")

            st.markdown("#### 📅 Cronograma de Intervenciones (Gantt)")
            df_gantt = df_filtrado.dropna(subset=['Inicio_DT', 'Fin_DT']).copy()
            if not df_gantt.empty:
                df_gantt['Activo'] = df_gantt['Central/Ubicacion'] + " | " + df_gantt['Equipo']
                fig_gantt = px.timeline(
                    df_gantt, x_start="Inicio_DT", x_end="Fin_DT", y="Activo", 
                    color="Disponibilidad_Equipo", hover_name="Empresa",
                    hover_data={"Tipo_Mantenimiento": True, "Sector": True, "MW_Efectivos": True},
                    color_discrete_map={'F/S': '#d62728', 'E/S': '#2ca02c', 'NO ESPECIFICADO': '#7f7f7f'}
                )
                fig_gantt.update_yaxes(autorange="reversed")
                fig_gantt.update_layout(height=max(400, len(df_gantt['Activo'].unique()) * 30))
                st.plotly_chart(fig_gantt, use_container_width=True, key="graf_mant_gantt")
        else:
            st.warning("No hay datos de mantenimiento programado para mostrar.")

    # ---------------------------------------------------------
    # PESTAÑA 3: POTENCIA INDISPONIBLE DETALLADA (ÁREAS F/S)
    # ---------------------------------------------------------
    with t_indisponible:
        if not df_gen_fs.empty:
            df_editor = df_gen_fs[['Empresa', 'Central/Ubicacion', 'Equipo', 'MW_Efectivos']].drop_duplicates(subset=['Central/Ubicacion', 'Equipo']).reset_index(drop=True)
            df_editor.rename(columns={'MW_Efectivos': 'Potencia_Indisponible_MW'}, inplace=True)
            
            st.markdown("#### 📝 Matriz de Asignación de Potencia")
            df_editado = st.data_editor(df_editor, use_container_width=True, hide_index=True)
            if st.button("💾 Guardar Potencias Editadas", type="primary"):
                guardar_potencias_asignadas(df_editado)
                st.success("✅ ¡Potencias almacenadas!")

            st.markdown("---")
            st.markdown("#### 📈 Detalle Evolutivo de Indisponibilidad (Ajustado sin Doble Conteo)")
            st.caption("Nota: El área bajo la curva representa la Energía Horaria (MWh) indisponible de los equipos F/S.")
            
            df_grafica_base = df_gen_fs.dropna(subset=['Inicio_DT', 'Fin_DT']).copy()
            
            if not df_grafica_base.empty and df_grafica_base['MW_Efectivos'].sum() > 0:
                df_grafica_base['Central_Equipo'] = df_grafica_base['Central/Ubicacion'] + " - " + df_grafica_base['Equipo']
                min_dt, max_dt = df_grafica_base['Inicio_DT'].min(), df_grafica_base['Fin_DT'].max()
                
                if min_dt < max_dt:
                    time_grid = pd.date_range(start=min_dt, end=max_dt, freq='h')
                    series_list = []
                    
                    for t in time_grid:
                        mask = (df_grafica_base['Inicio_DT'] <= t) & (df_grafica_base['Fin_DT'] > t)
                        df_t_unique = df_grafica_base.loc[mask].drop_duplicates(subset=['Central_Equipo'])
                        
                        centrales_completas = df_t_unique[df_t_unique['Equipo'] == 'CENTRAL'][['Empresa', 'Central/Ubicacion']].drop_duplicates()
                        for _, row_c in centrales_completas.iterrows():
                            condicion_remover = (df_t_unique['Empresa'] == row_c['Empresa']) & (df_t_unique['Central/Ubicacion'] == row_c['Central/Ubicacion']) & (df_t_unique['Equipo'] != 'CENTRAL')
                            df_t_unique = df_t_unique[~condicion_remover]
                        
                        for _, row_valid in df_t_unique.iterrows():
                            series_list.append({
                                'Fecha_Hora': t,
                                'Central_Unidad': row_valid['Central_Equipo'],
                                'Energía_MWh': row_valid['MW_Efectivos']
                            })
                    
                    df_area_detallado = pd.DataFrame(series_list)
                    if not df_area_detallado.empty:
                        fig_area = px.area(
                            df_area_detallado, 
                            x='Fecha_Hora', 
                            y='Energía_MWh', 
                            color='Central_Unidad',
                            title="Curva Detallada de Potencia / Energía Indisponible F/S",
                            labels={'Energía_MWh': 'Energía Horaria (MWh) / Potencia (MW)', 'Fecha_Hora': 'Intervalo Operativo'}
                        )
                        fig_area.update_traces(line=dict(width=0))
                        fig_area.update_layout(hovermode="x unified", margin=dict(t=80))
                        st.plotly_chart(fig_area, use_container_width=True, key="graf_indisp_fs")
                else:
                    st.info("Intervalo de tiempo demasiado estrecho para construir la curva.")
        else:
            st.warning("No hay maniobras F/S para gestionar o graficar en este periodo.")

    # ---------------------------------------------------------
    # PESTAÑA 4: DESPACHO EXCEL (GAS) + ANÁLISIS DE RESERVAS
    # ---------------------------------------------------------
    with t_termico_gas:
        st.subheader("⚡ Despacho Programado/Reprogramado (Unidades de Gas)")
        
        dfs_tab_gas = []
        renombrado_gas = {cod: f"{info['central']} - {info['unidad']}" for cod, info in INFO_UNIDADES_GAS.items()}
        label_to_code_gas = {f"{info['central']} - {info['unidad']}": cod for cod, info in INFO_UNIDADES_GAS.items()}
        
        for f in fechas_ordenadas:
            if f not in active_prog_dict: continue
            active_prog = active_prog_dict[f]
            dics_cache = dics_cache_dict[f]
            
            dia_data = {}
            for i in range(48):
                p = active_prog[i]
                df_excel = dics_cache[p].get("EXCEL")
                if df_excel is not None:
                    for cod_col in df_excel.columns:
                        if cod_col == 'Hora': continue
                        if cod_col in renombrado_gas:
                            friendly_name = renombrado_gas[cod_col]
                            if friendly_name not in dia_data: dia_data[friendly_name] = [0.0]*48
                            val = df_excel[cod_col].iloc[i] if i < len(df_excel) else 0.0
                            try: dia_data[friendly_name][i] += float(val) if pd.notna(val) else 0.0
                            except: pass
                        
            if dia_data:
                df_dia = pd.DataFrame(dia_data)
                df_dia.insert(0, 'Hora', ts_dict[f])
                dfs_tab_gas.append(df_dia)
                
        if dfs_tab_gas:
            df_total_gas = pd.concat(dfs_tab_gas, ignore_index=True).fillna(0)
            num_cols = [c for c in df_total_gas.columns if c != 'Hora']
            active_cols = [c for c in num_cols if df_total_gas[c].sum() > 0]
            
            if active_cols:
                col_g1, col_g2 = st.columns([3, 1])
                todas_centrales = sorted(active_cols)
                
                with col_g1:
                    filtro_cen_g = st.multiselect("⚡ Filtrar Turbinas de Gas (Despacho):", options=todas_centrales, default=[], placeholder="Todas (vacío) o buscar...", key="filtro_cen_gas")
                with col_g2:
                    tipo_grafico_g = st.radio("Estilo de Gráfica:", ["Área Apilada (MW)", "Barras Apiladas (Energía MWh/Día)"], horizontal=True, key="graf_radio_gas")
                
                lista_filtro = filtro_cen_g if filtro_cen_g else todas_centrales
                df_plot = df_total_gas[['Hora'] + lista_filtro]
                tipo_str = "barra" if "Energía" in tipo_grafico_g else "area"
                
                fig_disp_gas = crear_grafica_dinamica(df_plot, marcadores=marcadores_globales, tipo_grafico=tipo_str, paleta_distinta=True)
                st.plotly_chart(fig_disp_gas, use_container_width=True, key="main_disp_gas")
                
                st.markdown("---")
                st.subheader("📉 Margen de Reserva Operativa (Gas)")
                
                col_r1, col_r2 = st.columns(2)
                centrales_disp_res = sorted(list(set([info['central'] for info in INFO_UNIDADES_GAS.values()])))
                sel_central_res = col_r1.multiselect("Filtro Central (Reserva Gas):", centrales_disp_res, default=[], key="cen_res_gas")
                
                unidades_disp_res = []
                for cod, info in INFO_UNIDADES_GAS.items():
                    if not sel_central_res or info['central'] in sel_central_res:
                        unidades_disp_res.append(renombrado_gas[cod])
                        
                sel_unidad_res = col_r2.multiselect("Filtro Unidad (Reserva Gas):", sorted(unidades_disp_res), default=[], key="uni_res_gas")
                cols_to_calc = sel_unidad_res if sel_unidad_res else unidades_disp_res
                
                if cols_to_calc:
                    df_reserva_gas = df_total_gas[['Hora']].copy()
                    for col_label in cols_to_calc:
                        cod = label_to_code_gas[col_label]
                        base_pot = dict_pot_editada.get(cod, INFO_UNIDADES_GAS[cod]['potencia'])
                        pot_series = pd.Series(base_pot, index=df_reserva_gas.index)
                        
                        if cod in dict_mant_vinculado:
                            vinculo = dict_mant_vinculado[cod]
                            parts = vinculo.split(" - ")
                            if len(parts) >= 2:
                                c_nom_v, u_nom_v = parts[0].strip().upper(), parts[1].strip().upper()
                                mask_fs = (df_gen_fs['Central/Ubicacion'].str.upper().str.contains(c_nom_v, na=False)) & \
                                          (df_gen_fs['Equipo'].str.upper().str.contains(u_nom_v, na=False))
                                eventos_vinculados = df_gen_fs[mask_fs]
                                for _, ev in eventos_vinculados.iterrows():
                                    ini, fin = ev['Inicio_DT'], ev['Fin_DT']
                                    if pd.notna(ini) and pd.notna(fin):
                                        in_fs = (df_reserva_gas['Hora'] >= ini) & (df_reserva_gas['Hora'] < fin)
                                        pot_series.loc[in_fs] = 0.0
                        
                        if cod in dict_mant_manual:
                            ini_m, fin_m = dict_mant_manual[cod]
                            in_fs_m = (df_reserva_gas['Hora'] >= ini_m) & (df_reserva_gas['Hora'] < fin_m)
                            pot_series.loc[in_fs_m] = 0.0
                        
                        despacho_real = df_total_gas[col_label].fillna(0) if col_label in df_total_gas.columns else 0.0
                        df_reserva_gas[col_label] = (pot_series - despacho_real).clip(lower=0)
                    
                    num_cols_res = [c for c in df_reserva_gas.columns if c != 'Hora']
                    df_reserva_gas['TOTAL_RESERVA'] = df_reserva_gas[num_cols_res].sum(axis=1)
                    df_res_melt = df_reserva_gas.melt(id_vars=['Hora', 'TOTAL_RESERVA'], var_name='Unidad_Ref', value_name='Reserva_MW')
                    color_map_res = get_stable_color_map(df_res_melt['Unidad_Ref'].unique())
                    
                    fig_res_gas = px.area(df_res_melt, x='Hora', y='Reserva_MW', color='Unidad_Ref', labels={'Reserva_MW': 'Reserva de Potencia (MW)'}, color_discrete_map=color_map_res)
                    fig_res_gas.add_scatter(x=df_reserva_gas['Hora'], y=df_reserva_gas['TOTAL_RESERVA'], mode='lines', line=dict(width=0, color='rgba(0,0,0,0)'), name='<b>∑ RESERVA TOTAL</b>', showlegend=False)
                    fig_res_gas.update_traces(line=dict(width=0))
                    
                    if marcadores_globales:
                        for ts, texto in marcadores_globales:
                            fig_res_gas.add_vline(x=ts, line_width=1.5, line_dash="dash", line_color="rgba(0,0,0,0.5)")
                            texto_con_hora = f"{texto.replace('(', '').replace(')', '')} {ts.strftime('%H:%M')}"
                            align = "left" if ts.hour == 0 and ts.minute == 30 else "center"
                            fig_res_gas.add_annotation(x=ts, y=1.08, yref="paper", text=f"<b>{texto_con_hora}</b>", showarrow=False, font=dict(size=10, color="white"), bgcolor="#e74c3c", textangle=-90, yanchor="bottom", xanchor=align)
                    
                    fig_res_gas.update_layout(hovermode="x unified", height=500, margin=dict(t=170, b=50, l=60, r=50))
                    st.plotly_chart(fig_res_gas, use_container_width=True, key="main_res_gas")
                else:
                    st.warning("No hay unidades seleccionadas para calcular la reserva operativa.")
            else:
                st.warning("No hay despacho de Gas reportado en este rango.")
        else:
            st.warning("No se pudo estructurar el despacho de Gas en este rango.")

    # ---------------------------------------------------------
    # PESTAÑA 5: DESPACHO EXCEL (DIÉSEL) + ANÁLISIS DE RESERVAS
    # ---------------------------------------------------------
    with t_termico_diesel:
        st.subheader("🛢️ Despacho Programado/Reprogramado (Unidades Diésel)")
        
        dfs_tab_d = []
        renombrado_diesel = {cod: f"{info['central']} - {info['unidad']}" for cod, info in INFO_UNIDADES_DIESEL.items()}
        label_to_code_diesel = {f"{info['central']} - {info['unidad']}": cod for cod, info in INFO_UNIDADES_DIESEL.items()}
        
        for f in fechas_ordenadas:
            if f not in active_prog_dict: continue
            active_prog = active_prog_dict[f]
            dics_cache = dics_cache_dict[f]
            
            dia_data = {}
            for i in range(48):
                p = active_prog[i]
                df_excel = dics_cache[p].get("EXCEL")
                if df_excel is not None:
                    for cod_col in df_excel.columns:
                        if cod_col == 'Hora': continue
                        if cod_col in renombrado_diesel:
                            friendly_name = renombrado_diesel[cod_col]
                            if friendly_name not in dia_data: dia_data[friendly_name] = [0.0]*48
                            val = df_excel[cod_col].iloc[i] if i < len(df_excel) else 0.0
                            try: dia_data[friendly_name][i] += float(val) if pd.notna(val) else 0.0
                            except: pass
                        
            if dia_data:
                df_dia = pd.DataFrame(dia_data)
                df_dia.insert(0, 'Hora', ts_dict[f])
                dfs_tab_d.append(df_dia)
                
        if dfs_tab_d:
            df_total_d = pd.concat(dfs_tab_d, ignore_index=True).fillna(0)
            num_cols = [c for c in df_total_d.columns if c != 'Hora']
            active_cols = [c for c in num_cols if df_total_d[c].sum() > 0]
            
            if active_cols:
                col_d1, col_d2 = st.columns([3, 1])
                todas_centrales_d = sorted(active_cols)
                
                with col_d1:
                    filtro_cen_d = st.multiselect("⚡ Filtrar Turbinas Diésel (Despacho):", options=todas_centrales_d, default=[], placeholder="Todas (vacío) o buscar...", key="filtro_cen_die")
                with col_d2:
                    tipo_grafico_d = st.radio("Estilo de Gráfica:", ["Área Apilada (MW)", "Barras Apiladas (Energía MWh/Día)"], horizontal=True, key="graf_radio_die")
                
                lista_filtro_d = filtro_cen_d if filtro_cen_d else todas_centrales_d
                df_plot_d = df_total_d[['Hora'] + lista_filtro_d]
                tipo_str_d = "barra" if "Energía" in tipo_grafico_d else "area"
                
                fig_disp_die = crear_grafica_dinamica(df_plot_d, marcadores=marcadores_globales, tipo_grafico=tipo_str_d, paleta_distinta=True)
                st.plotly_chart(fig_disp_die, use_container_width=True, key="main_disp_die")
                
                st.markdown("---")
                st.subheader("📉 Margen de Reserva Operativa (Diésel)")
                
                col_r1d, col_r2d = st.columns(2)
                centrales_disp_res_d = sorted(list(set([info['central'] for info in INFO_UNIDADES_DIESEL.values()])))
                sel_central_res_d = col_r1d.multiselect("Filtro Central (Reserva Diésel):", centrales_disp_res_d, default=[], key="cen_res_die")
                
                unidades_disp_res_d = []
                for cod, info in INFO_UNIDADES_DIESEL.items():
                    if not sel_central_res_d or info['central'] in sel_central_res_d:
                        unidades_disp_res_d.append(renombrado_diesel[cod])
                        
                sel_unidad_res_d = col_r2d.multiselect("Filtro Unidad (Reserva Diésel):", sorted(unidades_disp_res_d), default=[], key="uni_res_die")
                cols_to_calc_d = sel_unidad_res_d if sel_unidad_res_d else unidades_disp_res_d
                
                if cols_to_calc_d:
                    df_reserva_die = df_total_d[['Hora']].copy()
                    for col_label in cols_to_calc_d:
                        cod = label_to_code_diesel[col_label]
                        base_pot = dict_pot_editada_d.get(cod, INFO_UNIDADES_DIESEL[cod]['potencia'])
                        pot_series = pd.Series(base_pot, index=df_reserva_die.index)
                        
                        if cod in dict_mant_vinc_d:
                            vinculo = dict_mant_vinc_d[cod]
                            parts = vinculo.split(" - ")
                            if len(parts) >= 2:
                                c_nom_v, u_nom_v = parts[0].strip().upper(), parts[1].strip().upper()
                                mask_fs = (df_gen_fs['Central/Ubicacion'].str.upper().str.contains(c_nom_v, na=False)) & \
                                          (df_gen_fs['Equipo'].str.upper().str.contains(u_nom_v, na=False))
                                eventos_vinculados = df_gen_fs[mask_fs]
                                for _, ev in eventos_vinculados.iterrows():
                                    ini, fin = ev['Inicio_DT'], ev['Fin_DT']
                                    if pd.notna(ini) and pd.notna(fin):
                                        in_fs = (df_reserva_die['Hora'] >= ini) & (df_reserva_die['Hora'] < fin)
                                        pot_series.loc[in_fs] = 0.0
                        
                        if cod in dict_mant_man_d:
                            ini_m, fin_m = dict_mant_man_d[cod]
                            in_fs_m = (df_reserva_die['Hora'] >= ini_m) & (df_reserva_die['Hora'] < fin_m)
                            pot_series.loc[in_fs_m] = 0.0
                        
                        despacho_real = df_total_d[col_label].fillna(0) if col_label in df_total_d.columns else 0.0
                        df_reserva_die[col_label] = (pot_series - despacho_real).clip(lower=0)
                    
                    num_cols_res_d = [c for c in df_reserva_die.columns if c != 'Hora']
                    df_reserva_die['TOTAL_RESERVA'] = df_reserva_die[num_cols_res_d].sum(axis=1)
                    df_res_melt_d = df_reserva_die.melt(id_vars=['Hora', 'TOTAL_RESERVA'], var_name='Unidad_Ref', value_name='Reserva_MW')
                    color_map_res_d = get_stable_color_map(df_res_melt_d['Unidad_Ref'].unique())
                    
                    fig_res_die = px.area(df_res_melt_d, x='Hora', y='Reserva_MW', color='Unidad_Ref', labels={'Reserva_MW': 'Reserva de Potencia (MW)'}, color_discrete_map=color_map_res_d)
                    fig_res_die.add_scatter(x=df_reserva_die['Hora'], y=df_reserva_die['TOTAL_RESERVA'], mode='lines', line=dict(width=0, color='rgba(0,0,0,0)'), name='<b>∑ RESERVA TOTAL</b>', showlegend=False)
                    fig_res_die.update_traces(line=dict(width=0))
                    
                    if marcadores_globales:
                        for ts, texto in marcadores_globales:
                            fig_res_die.add_vline(x=ts, line_width=1.5, line_dash="dash", line_color="rgba(0,0,0,0.5)")
                            texto_con_hora = f"{texto.replace('(', '').replace(')', '')} {ts.strftime('%H:%M')}"
                            align = "left" if ts.hour == 0 and ts.minute == 30 else "center"
                            fig_res_die.add_annotation(x=ts, y=1.08, yref="paper", text=f"<b>{texto_con_hora}</b>", showarrow=False, font=dict(size=10, color="white"), bgcolor="#e74c3c", textangle=-90, yanchor="bottom", xanchor=align)
                    
                    fig_res_die.update_layout(hovermode="x unified", height=500, margin=dict(t=170, b=50, l=60, r=50))
                    st.plotly_chart(fig_res_die, use_container_width=True, key="main_res_die")
                else:
                    st.warning("No hay unidades seleccionadas para calcular la reserva operativa.")
            else:
                st.warning("No hay despacho de Diésel reportado en este rango.")
        else:
            st.warning("No se pudo estructurar el despacho Diésel en este rango.")

    # ---------------------------------------------------------
    # PESTAÑA 1 (PARTE B): GRÁFICAS DE RESUMEN GLOBAL 
    # ---------------------------------------------------------
    with t_resumen_global:
        st.markdown("---")
        st.subheader("📈 Curva de Reserva Operativa Total del Sistema (Gas + Diésel)")
        
        if not df_reserva_gas.empty or not df_reserva_die.empty:
            hora_col = df_reserva_gas['Hora'] if not df_reserva_gas.empty else df_reserva_die['Hora']
            res_gas_sum = df_reserva_gas['TOTAL_RESERVA'] if not df_reserva_gas.empty else 0
            res_die_sum = df_reserva_die['TOTAL_RESERVA'] if not df_reserva_die.empty else 0
            
            df_res_total = pd.DataFrame({
                'Hora': hora_col,
                'Reserva Térmica - Gas (MW)': res_gas_sum,
                'Reserva Térmica - Diésel (MW)': res_die_sum,
                'Reserva Total (MW)': res_gas_sum + res_die_sum
            })
            
            fig_sum = px.area(
                df_res_total, x='Hora', y=['Reserva Térmica - Gas (MW)', 'Reserva Térmica - Diésel (MW)'], 
                labels={'value': 'Potencia Reservada (MW)', 'variable': 'Tecnología'},
                color_discrete_map={'Reserva Térmica - Gas (MW)': '#27ae60', 'Reserva Térmica - Diésel (MW)': '#c0392b'}
            )
            fig_sum.add_scatter(x=df_res_total['Hora'], y=df_res_total['Reserva Total (MW)'], mode='lines', 
                                line=dict(width=0, color='rgba(0,0,0,0)'), name='<b>∑ RESERVA TOTAL</b>', showlegend=False)
            fig_sum.update_traces(line=dict(width=0))
            if marcadores_globales:
                for ts, texto in marcadores_globales:
                    fig_sum.add_vline(x=ts, line_width=1.5, line_dash="dash", line_color="rgba(0,0,0,0.5)")
                    texto_con_hora = f"{texto.replace('(', '').replace(')', '')} {ts.strftime('%H:%M')}"
                    align = "left" if ts.hour == 0 and ts.minute == 30 else "center"
                    fig_sum.add_annotation(x=ts, y=1.08, yref="paper", text=f"<b>{texto_con_hora}</b>", showarrow=False, font=dict(size=10, color="white"), bgcolor="#2980b9", textangle=-90, yanchor="bottom", xanchor=align)
            
            fig_sum.update_layout(hovermode="x unified", height=500, margin=dict(t=170, b=50, l=60, r=50))
            st.plotly_chart(fig_sum, use_container_width=True, key="graf_resumen_suma_global")
            
            min_res_tot = df_res_total['Reserva Total (MW)'].min()
            max_res_tot = df_res_total['Reserva Total (MW)'].max()
            hora_min_tot = df_res_total.loc[df_res_total['Reserva Total (MW)'].idxmin(), 'Hora'].strftime('%d/%m/%Y %H:%M')
            hora_max_tot = df_res_total.loc[df_res_total['Reserva Total (MW)'].idxmax(), 'Hora'].strftime('%d/%m/%Y %H:%M')
            
            st.success(f"🛡️ **Reserva Operativa Global (Gas + Diésel):** Máxima de **{max_res_tot:,.2f} MW** (el {hora_max_tot}) | Mínima crítica de **{min_res_tot:,.2f} MW** (el {hora_min_tot})")
        else:
            st.warning("⚠️ Asegúrate de asignar las potencias en las tablas superiores para generar la curva combinada.")
            
        st.markdown("---")
        
        # Despliegue Vertical de Gráficas
        st.markdown("#### ⚡ Panorama Operativo (Gas)")
        if fig_disp_gas: 
            st.plotly_chart(fig_disp_gas, use_container_width=True, key="graf_resumen_disp_gas")
            df_disp_gas_cols = [c for c in df_total_gas.columns if c != 'Hora']
            disp_gas_sum = df_total_gas[df_disp_gas_cols].sum(axis=1)
            min_disp_g, max_disp_g = disp_gas_sum.min(), disp_gas_sum.max()
            h_min_dg, h_max_dg = df_total_gas.loc[disp_gas_sum.idxmin(), 'Hora'].strftime('%d/%m/%Y %H:%M'), df_total_gas.loc[disp_gas_sum.idxmax(), 'Hora'].strftime('%d/%m/%Y %H:%M')
            st.info(f"📊 **Despacho Gas:** Máximo despacho de **{max_disp_g:,.2f} MW** (el {h_max_dg}) | Mínimo despacho de **{min_disp_g:,.2f} MW** (el {h_min_dg})")
        
        if fig_res_gas: 
            st.plotly_chart(fig_res_gas, use_container_width=True, key="graf_resumen_res_gas")
            min_res_gas = df_reserva_gas['TOTAL_RESERVA'].min()
            max_res_gas = df_reserva_gas['TOTAL_RESERVA'].max()
            hora_min_gas = df_reserva_gas.loc[df_reserva_gas['TOTAL_RESERVA'].idxmin(), 'Hora'].strftime('%d/%m/%Y %H:%M')
            hora_max_gas = df_reserva_gas.loc[df_reserva_gas['TOTAL_RESERVA'].idxmax(), 'Hora'].strftime('%d/%m/%Y %H:%M')
            st.info(f"🛡️ **Reserva Operativa Gas:** Máxima reserva de **{max_res_gas:,.2f} MW** (el {hora_max_gas}) | Mínima reserva de **{min_res_gas:,.2f} MW** (el {hora_min_gas})")
            
        st.markdown("---")
        st.markdown("#### 🛢️ Panorama Operativo (Diésel / Residual)")
        if fig_disp_die: 
            st.plotly_chart(fig_disp_die, use_container_width=True, key="graf_resumen_disp_die")
            df_disp_d_cols = [c for c in df_total_d.columns if c != 'Hora']
            disp_d_sum = df_total_d[df_disp_d_cols].sum(axis=1)
            min_disp_d, max_disp_d = disp_d_sum.min(), disp_d_sum.max()
            h_min_dd, h_max_dd = df_total_d.loc[disp_d_sum.idxmin(), 'Hora'].strftime('%d/%m/%Y %H:%M'), df_total_d.loc[disp_d_sum.idxmax(), 'Hora'].strftime('%d/%m/%Y %H:%M')
            st.info(f"📊 **Despacho Diésel:** Máximo despacho de **{max_disp_d:,.2f} MW** (el {h_max_dd}) | Mínimo despacho de **{min_disp_d:,.2f} MW** (el {h_min_dd})")
        
        if fig_res_die: 
            st.plotly_chart(fig_res_die, use_container_width=True, key="graf_resumen_res_die")
            min_res_die = df_reserva_die['TOTAL_RESERVA'].min()
            max_res_die = df_reserva_die['TOTAL_RESERVA'].max()
            hora_min_die = df_reserva_die.loc[df_reserva_die['TOTAL_RESERVA'].idxmin(), 'Hora'].strftime('%d/%m/%Y %H:%M')
            hora_max_die = df_reserva_die.loc[df_reserva_die['TOTAL_RESERVA'].idxmax(), 'Hora'].strftime('%d/%m/%Y %H:%M')
            st.info(f"🛡️ **Reserva Operativa Diésel:** Máxima reserva de **{max_res_die:,.2f} MW** (el {hora_max_die}) | Mínima reserva de **{min_res_die:,.2f} MW** (el {hora_min_die})")

    # ---------------------------------------------------------
    # PESTAÑA 6: MOTIVOS DE REPROGRAMAS (1 a 1)
    # ---------------------------------------------------------
    with t_motivos:
        st.subheader("📋 Motivos de Reprogramación Operativa (RDO)")
        tabla_motivos = []
        for f in fechas_ordenadas:
            if f not in data_yupana: continue
            df_dia_sel = data_yupana[f]["Dataframes"]
            progs = [p for p in ["PDO"] + [f"RDO_{l}" for l in rdo_letras] if p in df_dia_sel]
            
            for p in progs:
                motivo_texto = df_dia_sel.get(f"MOTIVO_{p}", "No se encontró justificación.")
                hora_inicio = df_dia_sel.get(f"HORA_{p}", "N/A")
                tabla_motivos.append({
                    "Fecha": f.strftime("%d/%m/%Y"),
                    "Reprograma": p,
                    "Hora de Inicio": hora_inicio,
                    "Justificación / Motivo": motivo_texto
                })
                
        if tabla_motivos:
            st.dataframe(pd.DataFrame(tabla_motivos), use_container_width=True)
        else:
            st.success("✅ No hubo reprogramas (RDO) extraídos en las fechas consultadas.")

    # ---------------------------------------------------------
    # PESTAÑA 7: TRAZABILIDAD
    # ---------------------------------------------------------
    with t_datos:
        st.dataframe(df_filtrado.drop(columns=['Potencia_Indisponible_MW'], errors='ignore') if not df_filtrado.empty else pd.DataFrame(), use_container_width=True)